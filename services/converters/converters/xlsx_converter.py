from typing import Dict, Any, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from datetime import datetime
from ..base_converter import BaseDocumentConverter, ConversionContext
from ..document_structure import DocumentStructure, DocumentElement, ElementType
from ..file_utils import FileUtils
from models.file_conversion_models import FileMetadata, FileType
from loguru import logger

class XlsxConverter(BaseDocumentConverter):
    """Enhanced XLSX to Markdown converter"""
    
    def __init__(self):
        super().__init__()
        self.structure = DocumentStructure()
        self.file_utils = FileUtils()
        self._temp_files = []
    
    async def convert(self, content: bytes, context: ConversionContext) -> Tuple[str, FileMetadata]:
        """Convert Excel file to markdown"""
        self.context = context
        temp_path = None
        
        try:
            # Create temporary file
            temp_path = self.file_utils.create_temp_file(content, '.xlsx')
            self._temp_files.append(temp_path)
            
            # Load workbook with data_only=True to get values instead of formulas
            wb = load_workbook(filename=temp_path, data_only=True, read_only=True)
            
            # Add workbook metadata
            self._add_workbook_metadata(wb)
            
            # Process each sheet
            total_tables = 0
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                if self._has_content(sheet):
                    self._process_sheet(sheet)
                    total_tables += 1
            
            # Convert to markdown
            markdown_content = self.structure.to_markdown()
            
            # Create metadata
            metadata = FileMetadata(
                filename=context.filename,
                size_bytes=context.size_bytes,
                file_type=FileType.XLSX,
                pages=len(wb.sheetnames),
                tables_count=total_tables,
                images_count=0,
                equations_count=0
            )
            
            return markdown_content, metadata
            
        except Exception as e:
            logger.error(f"XLSX conversion error: {str(e)}")
            raise
        finally:
            # Cleanup temporary files
            self.file_utils.cleanup_temp_files(self._temp_files)
            if wb:
                wb.close()
    
    def _add_workbook_metadata(self, wb) -> None:
        """Add workbook metadata to document structure"""
        try:
            metadata = {
                'sheets': wb.sheetnames,
                'properties': {
                    'title': wb.properties.title if hasattr(wb.properties, 'title') else None,
                    'subject': wb.properties.subject if hasattr(wb.properties, 'subject') else None,
                    'creator': wb.properties.creator if hasattr(wb.properties, 'creator') else None,
                    'created': wb.properties.created.isoformat() if hasattr(wb.properties, 'created') and wb.properties.created else None,
                    'modified': wb.properties.modified.isoformat() if hasattr(wb.properties, 'modified') and wb.properties.modified else None
                }
            }
            
            self.structure.add_element(DocumentElement(
                type=ElementType.METADATA,
                content=metadata,
                metadata={'source': 'workbook_properties'}
            ))
            
        except Exception as e:
            self.log_warning(f"Error extracting workbook metadata: {str(e)}")
    
    def _has_content(self, sheet: Worksheet) -> bool:
        """Check if sheet has any content"""
        try:
            # In read_only mode, we need to iterate
            for row in sheet.iter_rows(min_row=1, max_row=1):
                return any(cell.value is not None for cell in row)
            return False
        except Exception:
            return False
    
    def _process_sheet(self, sheet: Worksheet) -> None:
        """Process single worksheet"""
        try:
            # Add sheet header
            self.structure.add_element(DocumentElement(
                type=ElementType.HEADING,
                content=f"Sheet: {sheet.title}",
                metadata={'sheet_name': sheet.title},
                level=2
            ))
            
            # Initialize table data
            table_data = []
            has_content = False
            max_cols = 0
            
            # Process rows
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if any(cell is not None for cell in row):
                    has_content = True
                    # Filter out empty trailing cells
                    while row and row[-1] is None:
                        row = row[:-1]
                    if row:
                        max_cols = max(max_cols, len(row))
                        # Format row data
                        formatted_row = [self._format_cell_value(cell) for cell in row]
                        table_data.append(formatted_row)
            
            # Ensure all rows have same number of columns
            for row in table_data:
                while len(row) < max_cols:
                    row.append('')
            
            # Add table to structure if content exists
            if has_content and table_data:
                self.structure.add_element(DocumentElement(
                    type=ElementType.TABLE,
                    content=table_data,
                    metadata={
                        'has_headers': True,
                        'align': ['left'] * max_cols,
                        'sheet': sheet.title
                    }
                ))
                
        except Exception as e:
            self.log_warning(f"Error processing sheet {sheet.title}: {str(e)}")
    
    def _format_cell_value(self, value: Any) -> str:
        """Format cell value for markdown"""
        if value is None:
            return ''
            
        if isinstance(value, datetime):
            return value.isoformat()
            
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)
            
        return str(value)
    
    def _format_column_width(self, width: Any) -> Optional[int]:
        """Convert Excel column width to character count"""
        try:
            if width is not None:
                # Excel's default font is calibri 11pt where 1 char ≈ 7 pixels
                # Standard width in Excel is 8.43 chars
                width_in_chars = float(width)
                return max(3, min(int(width_in_chars), 50))  # Limit between 3-50 chars
        except (ValueError, TypeError):
            pass
        return None